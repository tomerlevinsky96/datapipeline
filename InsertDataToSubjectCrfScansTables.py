import numpy as np
import gspread
from google.oauth2.service_account import Credentials
from psycopg2 import sql
import pandas as pd
import psycopg2
from openpyxl.workbook import Workbook
from psycopg2 import Error
import re
import os
import openpyxl
import create_scan_type_table
from create_scan_type_table import get_df
# Function to connect to the PostgreSQL database
import random
import string
def connect_to_db():
    try:
        connection = psycopg2.connect(
            user="tomer",
            password="t1",
            host="localhost",
            port="5433",  # Assuming default PostgreSQL port
            database="postgres5",
        )
        connection.autocommit = True
        return connection
    except Error as e:
        print("Error while connecting to PostgreSQL:", e)

source_db_config = {
    'user': 'tomer',
    'password': 't1',
    'host': 'localhost',
    'port': '5433',
    'database': 'postgres5'
}

# Connection parameters for destination database
destination_db_config = {
    'user': 'tomer',
    'password': 't1',
    'host': 'localhost',
    'port': '5433',
    'database': 'appdb'
}

def truncate_all_tables(DB_CONFIG):
    try:
        # Connect to PostgreSQL database
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Fetch all table names in the public schema
        cursor.execute("""
            SELECT tablename 
            FROM pg_tables 
            WHERE schemaname = 'public';
        """)
        tables = cursor.fetchall()

        # Generate and execute TRUNCATE command for all tables
        if tables:
            truncate_query = sql.SQL("TRUNCATE TABLE {} RESTART IDENTITY CASCADE").format(
                sql.SQL(', ').join(sql.Identifier(table[0]) for table in tables)
            )
            cursor.execute(truncate_query)
            conn.commit()
            print("All tables truncated successfully.")
        else:
            print("No tables found in the public schema.")

    except (Exception, psycopg2.DatabaseError) as error:
        print(f"Error occurred: {error}")
    finally:
        # Close the cursor and connection
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def generate_random_code(length=10):
    """Generate a random alphanumeric code of specified length."""
    characters = string.ascii_letters + string.digits  # Includes letters and digits
    code = ''.join(random.choice(characters) for _ in range(length))
    return code

def findGuid(ID_or_phone,cursor,connection):
    try:
     inser_query=f"""select guid from crf where id='{ID_or_phone}'"""
     cursor.execute(inser_query)
     guid=cursor.fetchall()
    except psycopg2.Error as e:
      print(e)
      connection.rollback()
    return guid


def get_columns(table_name, cursor):
    query_command = f"""SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'"""
    cursor.execute(query_command)
    columns = [row[0] for row in cursor.fetchall()]
    return columns

def get_columns_ordinal_postion(table_name, cursor):
    query_command = f"""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = '{table_name}'
        ORDER BY ordinal_position
    """
    cursor.execute(query_command)
    columns = [row[0] for row in cursor.fetchall()]
    return columns
def reorder_rows_by_ID(Crf_Data):
    Crf_Data['ID'] = pd.to_numeric(Crf_Data['ID'], errors='coerce')

    # Sort the DataFrame, placing NaN values at the end
    Crf_Data = Crf_Data.sort_values(by='ID', na_position='last')

    # Convert 'ID' to string, handling NaN values explicitly
    Crf_Data['ID'] = Crf_Data['ID'].apply(lambda x: str(int(x)) if pd.notna(x) else '')

    # Reset index if needed
    Crf_Data.reset_index(drop=True, inplace=True)
    return Crf_Data

def build_query_with_values(tablename, colnames, row):
    # Format column names and values for the insert query
    colnames_str = ", ".join(colnames)
    values_str = ", ".join([f"'{str(value)}'" for value in row])
    query = f"INSERT INTO {tablename} ({colnames_str}) VALUES ({values_str})"
    return query

def process_data_table(tablename,excluded_columns):

    try:
        # Connect to the source database
        source_conn = psycopg2.connect(**source_db_config)
        source_cursor = source_conn.cursor()
        dest_conn = psycopg2.connect(**destination_db_config)
        dest_cursor = dest_conn.cursor()
        # Get all column names
        all_columns = get_columns_ordinal_postion(tablename, source_cursor)

        # Exclude specified columns
        colnames = [col for col in all_columns if col not in excluded_columns]


        # Fetch data from source database
        source_cursor.execute(sql.SQL("SELECT {} FROM {}").format(
            sql.SQL(', ').join(map(sql.Identifier, [col for col in all_columns if col not in excluded_columns])),
            sql.Identifier(tablename)
        ))
        data = source_cursor.fetchall()
    except Exception as e:
        print(f"Error processing CRF table: {e}")
    insert_data_into_destination(tablename, data, colnames)


# Function to process and insert data from CrfYaData DataFrame into the Patients table
def Crf_to_subjet_table(Crf_Data,cursor, connection):
    Crf_Data=reorder_rows_by_ID(Crf_Data)
    for index, row in Crf_Data.iterrows():
        try:
            questionaire_code = row['Qcode']
            if str(questionaire_code) != 'nan':
              if row['ID']!='' :
                  select_query = f"""select guid from subjects where id='{row['ID']}'"""
                  cursor.execute(select_query)
                  guid = cursor.fetchone()
                  if not guid:
                      while True:
                          guid = generate_random_code(length=9)
                          select_query = f"""select guid from subjects where guid='{guid}'"""
                          cursor.execute(select_query)
                          already_exsisting_guid = cursor.fetchall()
                          if not already_exsisting_guid:
                              insert_query = f"INSERT INTO subjects(GuId,QuestionaireCode,id,email) VALUES ('{guid}','{questionaire_code}','{row['ID']}','{row['Email']}');"
                              cursor.execute(insert_query)
                              connection.commit()
                              break

                  else:
                      insert_query = f"INSERT INTO subjects(GuId,QuestionaireCode,id,email) VALUES ('{guid[0]}','{questionaire_code}','{row['ID']}','{row['Email']}');"
                      cursor.execute(insert_query)
                      connection.commit()
              elif row['Email']!='' :
                  select_query = f"""select guid from subjects where email='{row['Email']}'"""
                  cursor.execute(select_query)
                  guid = cursor.fetchone()
                  if guid:
                      insert_query = f"INSERT INTO subjects(GuId,QuestionaireCode,id,email) VALUES ('{guid}','{questionaire_code}','{row['ID']}','{row['Email']}');"
                      cursor.execute(insert_query)
                      connection.commit()
        except psycopg2.Error as e:
            print(e)
            connection.rollback()

def fetch_data_from_source(tablename, exclude_columns):
    try:
        # Connect to the source database
        source_conn = psycopg2.connect(**source_db_config)
        source_cursor = source_conn.cursor()

        # Get all column names
        all_columns = get_columns(tablename, source_cursor)

        # Exclude specified columns
        colnames = [col for col in all_columns if col not in exclude_columns]

        # Fetch data from source database
        source_cursor.execute(sql.SQL("SELECT {} FROM {}").format(
            sql.SQL(', ').join(map(sql.Identifier, colnames)),
            sql.Identifier(tablename)
        ))
        data = source_cursor.fetchall()

        # Close connection
        source_cursor.close()
        source_conn.close()

        return data, colnames
    except Exception as e:
        print(f"Error fetching data from source: {e}")
        return None, None

def insert_data_into_destination(tablename, data, colnames):
    try:
        # Connect to the destination database
        dest_conn = psycopg2.connect(**destination_db_config)
        dest_cursor = dest_conn.cursor()

        # Iterate over each row in the data
        for row in data:
            # Build the insert query with actual values for each row
            try:
             insert_query = build_query_with_values(tablename, colnames, row)
             # Print the query before execution
             print(f"Executing query: {insert_query}")
             dest_cursor.execute(insert_query)
             dest_conn.commit()
            except Exception as e:
             print(f"Error inserting data into destination: {e}")
             dest_conn.rollback()
        # Commit changes
        dest_conn.commit()

        # Close connection
        dest_cursor.close()
        dest_conn.close()
    except Exception as e:
        print(f"Error inserting data into destination: {e}")


# Function to transform and insert data from CrfYaData DataFrame into the CRF table
def Crf_Data_crf_table(Crf_Data, cursor, connection):
    Crf_Data = reorder_rows_by_ID(Crf_Data)
    # Find the index of the 'Name' column
    name_index = Crf_Data.columns.get_loc('Name')
    # Split the 'Name' column into 'FirstName' and 'LastName'
    Crf_Data[['FirstName', 'LastName']] = Crf_Data['Name'].str.split(' ', n=1, expand=True)
    # Drop the original 'Name' column
    Crf_Data.drop(columns=['Name'], inplace=True)
    # Insert 'FirstName' and 'LastName' columns at the same location as the original 'Name' column
    Crf_Data.insert(name_index, 'FirstName', Crf_Data.pop('FirstName'))
    Crf_Data.insert(name_index + 1, 'LastName', Crf_Data.pop('LastName'))
    for index, row in Crf_Data.iterrows():
        if str(row['ScanID']) != 'NaT' and str(row['ScanID']) !='':
            if row['ID']!='':
                  select_query = f"""select guid from subjects where ID='{row['ID']}'"""
                  cursor.execute(select_query)
                  guid = cursor.fetchone()
                  if guid:
                     values = [guid[0]] + [str(value) for value in row]
                     values = ', '.join(f"'{value}'" for value in values)
                     try:
                        insert_query = f"""
                           INSERT INTO CRF (
                             guid,datetimescan,status,lab,firstname,lastname,id,cellularno,email,gender,dob,scandate,ageofscan
                             ,weight,height,protocol,study,groupname,scantag,questionairecode,noscan,scandayqday) VALUES ({values});
                           """
                        cursor.execute(insert_query)
                        connection.commit()
                     except psycopg2.Error as e:
                        print(e)
            elif row['Email']!='':
              try:
                  select_query = f"""select guid from subjects where email='{row['Email']}'"""
                  cursor.execute(select_query)
                  guid = cursor.fetchone()
                  if guid:
                     insert_query = f"""
                                  INSERT INTO CRF (
                                    guid,scanid,status,lab,firstname,lastname,id,cellularno,email,gender,dob,scandate,ageofscan
                                    ,weight,height,protocol,study,groupname,scantag,questionairecode,noscan,scandayqday,datetimescan) VALUES ({values});
                              """
                     cursor.execute(insert_query)
                     connection.commit()
              except psycopg2.Error as e:
                print(e)
# Function to process and insert data from SnBBData DataFrame into the SharedScans table
def process_and_insert_SnBBData(SnBBData, cursor, connection):
    SnBBData['DateTimeScan'] = SnBBData['Path'].str.extract(r'(\d{8}_\d{4})')
    SnBBData['DateTimeScan'] = pd.to_datetime(SnBBData['DateTimeScan'], format='%Y%m%d_%H%M')
    cols = SnBBData.columns.tolist()
    cols.insert(0, cols.pop(cols.index('DateTimeScan')))
    SnBBData = SnBBData.reindex(columns=cols)

    for index, row in SnBBData.iterrows():
        if row['DateTimeScan'] != 'NaT':
            values = ', '.join(f"'{value}'" for value in row)
            parts = values.split(', ', 1)
            values = f"{parts[0]},NULL,NULL,{parts[1]},'SNBB'"
            try:
                insert_query = f"""
                    INSERT INTO Scans (
                        DateTimeScan,bidspath,resultspath,rawdatapath, T1w, T2w, Flair, RestFmri, TaskFmri, TaskNames, DmriAp, DmriPa, Mrtrix, AxSi, Qsi,
                        HcpFreesurfer, HcpRestFMRI, HcpTaskFmri, HcpDiffusion, DataLocation
                    ) VALUES ({values});
                """
                cursor.execute(insert_query)
                connection.commit()
            except psycopg2.Error as e:
                print(e)
                connection.rollback()

# Function to process and insert data from YaSharedScansData DataFrame into the SharedScans table
def process_and_insert_YaSharedScansData(YaSharedScansData, cursor, connection):
    YaSharedScansData['DateTimeScan'] = YaSharedScansData['Path'].str.extract(r'(\d{8}_\d{4})')
    YaSharedScansData['DateTimeScan'] = pd.to_datetime(YaSharedScansData['DateTimeScan'], format='%Y%m%d_%H%M')
    cols = YaSharedScansData.columns.tolist()
    cols.insert(0, cols.pop(cols.index('DateTimeScan')))
    YaSharedScansData = YaSharedScansData.reindex(columns=cols)

    for index, row in YaSharedScansData.iterrows():
        if row['DateTimeScan'] != pd.NaT:
            values = ', '.join(f"'{value}'" for value in row)
            parts = values.split(', ', 1)
            values = f"{parts[0]},NULL,NULL,{parts[1]},'ya_shared_scans'"
            try:
                insert_query = f"""
                    INSERT INTO Scans (
                        DateTimeScan,bidspath, resultspath,rawdatapath,T1w, T2w, Flair, RestFmri, TaskFmri, TaskNames, DmriAp, DmriPa, Mrtrix, AxSi, Qsi,
                        HcpFreesurfer, HcpRestFMRI, HcpTaskFmri, HcpDiffusion, DataLocation
                    ) VALUES ({values});
                """
                cursor.execute(insert_query)
                connection.commit()
            except psycopg2.Error as e:
                print(e)
                connection.rollback()



# Function to read the Excel and CSV files and insert data into the PostgreSQL database
def open_and_read_google(sheet_id):
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)
    workbook = client.open_by_key(sheet_id)
    # Get all values from the worksheet
    worksheet = workbook.get_worksheet(0)
    data=worksheet.get_all_values()
    data=pd.DataFrame(data[1:], columns=data[0])
    return data

def INSERT_INTO_QUESTIONAIRE_TABLE(questionaireVersion):
    connection = connect_to_db()
    if connection is None:
        return
    cursor = connection.cursor()
    for index, row in questionaireVersion.iterrows():
        try:
            insert_query = f"INSERT INTO questionaire (dateTime) VALUES ('{row['חותמת זמן']}');"
            cursor.execute(insert_query)
            connection.commit()
        except Error as e:
            print("Error inserting data:", e)


# Function to insert data into the questiones table
def INSERT_INTO_QUESTIONES_TABLE(Questions):
    connection = connect_to_db()
    if connection is None:
        return
    cursor = connection.cursor()
    columns = get_columns('questiones', cursor)

    # Retrieve existing questions from the questiones table
    select_query = "SELECT DISTINCT question FROM questiones"
    cursor.execute(select_query)
    results = cursor.fetchall()
    QuestionsfromTable = [row[0] for row in results]

    for value in Questions:
        if value not in QuestionsfromTable:
            value = value.replace("'", "''")
            try:
                insert_query = f"INSERT INTO questiones (Question) VALUES ('{value}');"
                cursor.execute(insert_query)
                connection.commit()
            except Error as e:
                print("Error inserting data:", e)


# Function to insert data into the answers table
def INSERT_INTO_ANSWERS_TABLE(Answers):
    connection = connect_to_db()
    if connection is None:
        return
    cursor = connection.cursor()

    for AnswerColumn in Answers.columns:
        if AnswerColumn != 'קוד הנבדק' and AnswerColumn != 'חותמת זמן':
            for index, row in Answers.iterrows():
                try:
                    # Sanitize the question column name
                    sanitized_column = AnswerColumn.replace("'", "''")
                    # Retrieve the QuestioneId from the questiones table
                    select_query = f"SELECT QuestioneId FROM questiones WHERE question='{sanitized_column}'"
                    cursor.execute(select_query)
                    QuestioneId = cursor.fetchone()[0]
                    # Prepare the answr and timestamp
                    Answer = row[AnswerColumn]
                    Answer = str(Answer).replace("'", "''")
                    # Retrieve the QuestionaireID from the questionaire table
                    select_query = f"SELECT QuestionaireID FROM questionaire WHERE dateTime='{row['חותמת זמן']}'"
                    cursor.execute(select_query)
                    QuestionaireID = cursor.fetchone()[0]
                    select_query=F"SELECT GuId FROM SUBJECTS WHERE QuestionaireCode='{row['קוד הנבדק']}'"
                    cursor.execute(select_query)
                    GuId = cursor.fetchone()
                    if GuId:
                      insert_query = f"INSERT INTO answers (GuId,QuestionaireCode, QuestioneId, QuestionaireID, Answer) VALUES ('{GuId[0]}','{row['קוד הנבדק']}', {QuestioneId}, {QuestionaireID}, N'{Answer}');"
                      cursor.execute(insert_query)
                      connection.commit()
                except Error as e:
                    if "syntax error" in str(e).lower():
                        print(f"Failed query: {insert_query}")
                    else:
                        continue



def read_excel_and_insert_data(Crf_Path,Qustionaire_path,Ya_Shared_Scans_path,SNBB_path):
    try:
        # Connect to the PostgreSQL database
        connection = connect_to_db()
        cursor = connection.cursor()
        Crf_Data = open_and_read_google(Crf_Path)
        questionaire_data = open_and_read_google(Qustionaire_path)
        SnBB_Scans_Data=pd.read_excel(SNBB_path)
        YaShared_Scans_Data=pd.read_excel(Ya_Shared_Scans_path)

        # SnBBData = create_scan_type_table.get_df(SnBBPath)
        # YaSharedScansData = create_scan_type_table.get_df(YaSharedScansPath)
        # snbb_excel_file_path = 'SnBBData.xlsx'
        # ya_shared_scans_excel_file_path = 'YaSharedScansData.xlsx'
        # SnBBData.to_excel(snbb_excel_file_path, sheet_name='SnBBData', index=False)
        # # Write the YaSharedScansData DataFrame to its own Excel file
        # YaSharedScansData.to_excel(ya_shared_scans_excel_file_path, sheet_name='YaSharedScansData', index=False)
        # print(f'Data successfully written to {snbb_excel_file_path} and {ya_shared_scans_excel_file_path}')
        # Strip whitespace from all columns in the dataframes
        columns_to_exclude = ['']
        Crf_Data = Crf_Data.drop(columns=columns_to_exclude, errors='ignore')
        for column in Crf_Data.columns:
            Crf_Data[column] = Crf_Data[column].apply(lambda x: str(x).strip())
        for column in SnBB_Scans_Data.columns:
            SnBB_Scans_Data[column] = SnBB_Scans_Data[column].apply(lambda x: str(x).strip())
        for column in YaShared_Scans_Data.columns:
           YaShared_Scans_Data[column] = YaShared_Scans_Data[column].apply(lambda x: str(x).strip())
        for column in questionaire_data.columns:
            questionaire_data[column] = [str(item).strip() for item in questionaire_data[column]]
        # Process and insert data into the PostgreSQL database
        Crf_to_subjet_table(Crf_Data,cursor, connection)
        Crf_Data_crf_table(Crf_Data, cursor, connection)
        process_and_insert_SnBBData(SnBB_Scans_Data, cursor, connection)
        process_and_insert_YaSharedScansData(YaShared_Scans_Data, cursor, connection)
        INSERT_INTO_QUESTIONAIRE_TABLE(questionaire_data)
        INSERT_INTO_QUESTIONES_TABLE(questionaire_data.columns.tolist())
        INSERT_INTO_ANSWERS_TABLE(questionaire_data)
        print("Data insertion complete!")
    except Error as e:
        print("Error while inserting data to PostgreSQL:", e)

def Insert_data_to_destination_database():
    tablenames = ['subjects', 'questionaire', 'questiones', 'crf', 'scans', 'answers']
    exclude_columns = ['id', 'phonenumber', 'email', 'firstname', 'lastname']
    for tablename in tablenames:
        if tablename == 'subjects':
            process_data_table(tablename, ['id', 'email'])
        elif tablename == 'crf':
            process_data_table('crf', ['dob', 'id', 'cellularno', 'email', 'firstname', 'lastname'])
        else:
            # Fetch data from source database
           data, colnames = fetch_data_from_source(tablename, exclude_columns)
           if data and colnames:
                #Insert data into destination database
                insert_data_into_destination(tablename, data, colnames)
           else:
                print(f"No data fetched from source for table {tablename}.")





def check_for_subject_with_missing_details(Questionaire_path):
    connection = connect_to_db()
    cursor = connection.cursor()
    selcect_query=f"""select subjectid from crf where id='nan'"""
    cursor.execute(selcect_query)
    # Fetch all the results
    Subjects_with_missing_data = cursor.fetchall()
    # Specify the path where the Excel file will be saved
    excel_path = "Subjects_with_missing_id.xlsx"
    # Save the Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the title for the sheet
    sheet.title = "Subjects rows with missing ID "
    row = 1
    # Add header (optional)
    sheet.cell(row=row, column=1).value = "Subject Codes"
    row += 1
    # Iterate through subject codes and add those not in subjectids to the Excel sheet
    for subject in subject_ids:
        sheet.cell(row=row, column=1).value = subject
        row += 1
    # Save the workbook
    workbook.save(excel_path)

    delete_query = f"DELETE FROM subjects WHERE subjectid IN ({subject_ids_str})"
    cursor.execute(delete_query)
    connection.commit()
    # Create a new workbook and get the active sheet
    print(f"Data has been successfully saved to {excel_path}")
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the title for the sheet
    sheet.title = "Questionaire Codes Not in CRF"

    # Fetch the subject IDs from the database
    select_query = "SELECT subjectid FROM subjects"
    cursor.execute(select_query)
    subjectids = [item[0] for item in cursor.fetchall()]  # Convert fetched tuples to a list of subject IDs
    # Read the Questionnaire data from Excel
    questionaire_data = pd.read_excel(Questionaire_path)
    subject_codes = questionaire_data['subject.code']
    # Initialize row counter for the Excel sheet
    row = 1
    # Add header (optional)
    sheet.cell(row=row, column=1).value = "Subject Codes"
    row += 1
    # Iterate through subject codes and add those not in subjectids to the Excel sheet
    for subject in subject_codes:
        if subject not in subjectids:
            sheet.cell(row=row, column=1).value = subject
            row += 1
    # Save the workbook
    workbook.save(excel_path)


# Paths to the files
CrfPath='1ZawIJ14Qep7r2Cs6PKI9lUVl5B2P3OKwdeTUARFQkag'
#Ya_Shared_Scans_path = '//132.66.46.62/Raw_Data'
Questionaire_path='1wyOBqgKe6mrSBQV32OAICFJIuWPqER_49cGaTvw41QM'
#Questionaire_path = '1wyOBqgKe6mrSBQV32OAICFJIuWPqER_49cGaTvw41QM'
#SNBB_path = '//132.66.46.165/snbb'
Ya_Shared_Scans_path ='YaSharedScansData.xlsx'
SNBB_path = 'SnBBData.xlsx'
truncate_all_tables(source_db_config)
truncate_all_tables(destination_db_config)
# Run the data pipeline
read_excel_and_insert_data(CrfPath,Questionaire_path,Ya_Shared_Scans_path,SNBB_path)
Insert_data_to_destination_database()

###check_for_subject_with_missing_details(Questionaire_path)
